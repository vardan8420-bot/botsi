"""
Report Generator - Сервис для создания отчетов (Excel, PDF)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from typing import List, Dict, Any


class ReportGeneratorService:
    """Сервис для генерации файлов отчетов"""
    
    def __init__(self):
        print("✅ Report Generator (Excel мастер) инициализирован")

    async def create_excel(self, filename: str, sheet_name: str, data: List[Dict[str, Any]]) -> str:
        """
        Создает красивый Excel файл
        
        Args:
            filename: Имя файла (без .xlsx)
            sheet_name: Имя листа
            data: Список словарей [{'Header1': 'Value1', 'Header2': 'Value2'}, ...]
        
        Returns:
            Путь к созданному файлу
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            if not data:
                return ""

            # 1. Заголовки
            headers = list(data[0].keys())
            
            # Стили для заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            centered_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = centered_alignment
                
                # Устанавливаем ширину колонки
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = 25

            # 2. Данные
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True)

            # Сохраняем
            path = f"{filename}.xlsx"
            wb.save(path)
            return path

        except Exception as e:
            print(f"❌ Ошибка создания Excel: {e}")
            return ""
